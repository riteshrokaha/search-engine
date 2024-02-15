from django.shortcuts import render
import string
from nltk.corpus import stopwords
from .forms import UploadFileForm
from .models import MyModel
from django.db.models import Q
import openpyxl
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.metrics import accuracy_score, precision_score, recall_score


# Create your views here.
def automate_data():
    excel_files = ['datasets.xlsx', 'publications.xlsx']
    for excel_file in excel_files:
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)  # Improve memory efficiency

            # Extract data from specific sheet (if not the first)
            sheet = wb['Sheet1']  # Consider user input or default based on your logic

            # Iterate through rows and cells, ensuring correct field mapping
            for row in sheet.iter_rows(min_row=2):  # Skip header row if present
                title = row[0].value
                publication_url = row[1].value
                author = row[2].value
                profile_url = row[3].value
                year = row[5].value
                # Process other fields
                obj, created = MyModel.objects.get_or_create(title=title, defaults={'author': author, 'year': year, 'publication_url': publication_url, 'profile_url': profile_url})  # Handle duplicates as needed

                if not created:
                    # Optionally update existing objects if allowed
                    obj.author = author
                    obj.year = year
                    obj.publication_url = publication_url
                    obj.profile_url = profile_url
                    obj.save()
        except Exception as e:
            print(e)



def home(request):
    if request.method == "POST":
        data = request.POST
        keyword = data.get("keyword")
        search_results = MyModel.objects.filter(Q(title__icontains=keyword) | Q(author__icontains=keyword) | Q(year__icontains=keyword))
        context = {
            "results": search_results,
            "search_keyword": keyword
        }
        return render(request, 'myscholar/home.html', context)
    
    try:
        automate_data()
    except:
        pass
    
    data = MyModel.objects.all()
    if data:
        context = {"data": data}
        return render(request, 'myscholar/home.html', context)

    return render(request, 'myscholar/home.html')


def upload_data(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            # Handle potential errors and edge cases (e.g., large files, invalid formats)
            if not excel_file.name.endswith('.xlsx'):
                return render(request, 'upload.html', {'form': form, 'error': 'Invalid file format. Please upload an XLSX file.'})
            
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)  # Improve memory efficiency

                # Extract data from specific sheet (if not the first)
                sheet = wb['Sheet1']  # Consider user input or default based on your logic

                # Iterate through rows and cells, ensuring correct field mapping
                for row in sheet.iter_rows(min_row=2):  # Skip header row if present
                    title = row[0].value
                    publication_url = row[1].value
                    author = row[2].value
                    profile_url = row[3].value
                    year = row[5].value
                    # Process other fields
                    obj, created = MyModel.objects.get_or_create(title=title, defaults={'author': author, 'year': year, 'publication_url': publication_url, 'profile_url': profile_url})  # Handle duplicates as needed

                    if not created:
                        # Optionally update existing objects if allowed
                        obj.author = author
                        obj.year = year
                        obj.publication_url = publication_url
                        obj.profile_url = profile_url
                        obj.save()

                return render(request, 'myscholar/upload.html', {'form': form, 'success': 'Data uploaded successfully!'})
            except Exception as e:
                return render(request, 'myscholar/upload.html', {'form': form, 'error': f'Error processing file: {e}'})
    else:
        form = UploadFileForm()

    return render(request, 'myscholar/upload.html', {'form': form})




def preprocess_text(text):
  # Lowercase conversion:
  text = text.lower()

  # Remove punctuation:
  for char in string.punctuation:
    text = text.replace(char, "")

  # Remove stop words:
  stop_words = stopwords.words('english')
  text = " ".join([word for word in text.split() if word not in stop_words])
  return text


def classification(request):
    if request.method == 'POST':
        data = request.POST
        sentence = preprocess_text(data.get("sentence"))

        # Load sentences and categories
        data = pd.read_excel("test data.xlsx")  

        # Train-Test split
        X_train, X_test, y_train, y_test = train_test_split(data["Sentence"], data["Category"], test_size=0.2, random_state=42)

        # Preprocess text data
        vectorizer = TfidfVectorizer()
        X_train_features = vectorizer.fit_transform(X_train)
        X_test_features = vectorizer.transform(X_test)

        # Train a Naive Bayes classifier
        model = MultinomialNB()
        model.fit(X_train_features, y_train)

        # Predict on test data
        y_pred = model.predict(X_test_features)

        # Evaluate model performance
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, average="weighted")
        recall = recall_score(y_test, y_pred, average="weighted")

        print("Accuracy:", accuracy)
        print("Precision:", precision)
        print("Recall:", recall)

        # Classify new sentences
        new_sentences = [sentence]
        new_features = vectorizer.transform(new_sentences)
        predictions = model.predict(new_features)

        print("\nClassifications:")
        for sentence, prediction in zip(new_sentences, predictions):
            print(f"- {sentence}: {prediction}")

        context = {
            'sentence': sentence,
            'prediction': prediction
        }

        return render(request, 'myscholar/classification.html', context)

    return render(request, 'myscholar/classification.html')
